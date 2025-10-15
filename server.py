from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime
import os


# Configuráveis pelo usuário
EXCEL_FILENAME = 'contabilidade.xlsx'
VALOR_ESTOQUE_INICIAL = 3000.00
CAPITAL_DE_GIRO = 950.00


app = Flask(__name__)


SHEETS = {
'vendas': 'Vendas',
'compras': 'Compras',
'despesas': 'Despesas'
}


def ensure_workbook(path):
"""Cria o arquivo Excel com abas e fórmulas padrão, se não existir."""
if os.path.exists(path):
wb = load_workbook(path)
return wb


wb = Workbook()
# Remove sheet default
default = wb.active
wb.remove(default)


# Criar abas de registro
for key in SHEETS.values():
ws = wb.create_sheet(key)
ws.append(['Data', 'Descrição', 'Valor (R$)'])
# largura colunas
ws.column_dimensions[get_column_letter(1)].width = 18
ws.column_dimensions[get_column_letter(2)].width = 40
ws.column_dimensions[get_column_letter(3)].width = 18
header_font = Font(bold=True)
for cell in ws[1]:
cell.font = header_font


# Aba de Config
cfg = wb.create_sheet('Config')
cfg.append(['Chave', 'Valor'])
cfg.append(['Estoque Inicial (R$)', VALOR_ESTOQUE_INICIAL])
app.run(host='0.0.0.0', port=5000, debug=True)